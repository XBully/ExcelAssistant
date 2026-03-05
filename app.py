import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import time
import base64
import xlrd
from xlutils.copy import copy as xl_copy

# --- 1. 页面配置与 CSS ---
st.set_page_config(page_title="小雷Excel批量助手", page_icon="⚡", layout="wide")

st.markdown("""
    <style>
    /* 全局背景与字体 */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
    
    .main {
        background-color: #fcfdfe;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }
    
    .block-container { 
        padding-top: 2rem !important; 
        max-width: 1000px !important;
    }
    
    header {visibility: hidden;}
    footer {visibility: hidden;}

    /* 标题样式 */
    .main-title {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 0.5rem;
        text-align: center;
        background: linear-gradient(90deg, #3b82f6, #06b6d4);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .sub-title {
        font-size: 0.9rem;
        color: #64748b;
        text-align: center;
        margin-bottom: 2rem;
    }

    /* 卡片通用样式 */
    .stFileUploader section {
        border: 2px dashed #e2e8f0 !important;
        background-color: #ffffff !important;
        padding: 1.5rem !important;
        border-radius: 12px !important;
        transition: all 0.3s ease;
    }
    .stFileUploader section:hover {
        border-color: #3b82f6 !important;
        background-color: #f8fafc !important;
    }

    .upload-card {
        background: #ffffff;
        border: 1px solid #f1f5f9;
        padding: 20px;
        border-radius: 16px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03);
        margin-bottom: 20px;
    }

    .config-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        padding: 18px;
        border-radius: 12px;
        margin-bottom: 15px;
    }

    .personal-box {
        padding: 15px;
        background: #ffffff;
        border-radius: 12px;
        margin-top: 10px;
        border: 1px solid #f1f5f9;
        box-shadow: inset 0 2px 4px 0 rgba(0, 0, 0, 0.02);
        border-left: 4px solid #3b82f6;
    }

    .download-bar {
        background: #ecfdf5;
        border: 1px solid #10b981;
        padding: 15px 20px;
        border-radius: 12px;
        margin: 20px 0;
        display: flex;
        align-items: center;
        justify-content: space-between;
        color: #065f46;
        font-weight: 500;
    }

    /* 按钮美化 */
    .stButton > button {
        border-radius: 10px !important;
        padding: 0.5rem 1rem !important;
        font-weight: 600 !important;
        transition: all 0.2s ease !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.25);
    }
    
    /* 侧边栏/输入框样式 */
    h5 {
        margin-bottom: 0.8rem !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        color: #1e293b !important;
        display: flex;
        align-items: center;
        gap: 8px;
    }
    
    .stExpander {
        border-radius: 12px !important;
        border: 1px solid #f1f5f9 !important;
        background: #ffffff !important;
    }
    </style>
    """, unsafe_allow_html=True)

if 'batch_results' not in st.session_state:
    st.session_state.batch_results = []
if 'extract_results' not in st.session_state:
    st.session_state.extract_results = []

# --- 2. 辅助函数 ---
def to_xlsx_stream(file):
    try:
        file.seek(0)
        df = pd.read_excel(file, header=None, engine='xlrd')
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False)
        output.seek(0)
        return output
    except: return None

def clean_columns(df):
    if df is None: return None
    new_cols, seen = [], {}
    
    for i, col in enumerate(df.columns):
        # 提取非 Unnamed 的部分，并清理空格
        parts = []
        if isinstance(col, tuple):
            # 处理多级表头：过滤掉空值、Unnamed、以及只含空格的字符串
            for p in col:
                p_str = str(p).strip()
                if p_str and "nan" not in p_str.lower() and "unnamed" not in p_str.lower():
                    parts.append(p_str)
        else:
            p_str = str(col).strip()
            if p_str and "nan" not in p_str.lower() and "unnamed" not in p_str.lower():
                parts = [p_str]
        
        # 组合名称，如果为空则设为 未命名
        name = " - ".join(parts) if parts else f"未命名_{i}"
        
        # 处理重名
        if name in seen:
            seen[name] += 1
            final_name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
            final_name = name
            
        new_cols.append(final_name)
    
    # 直接赋值，不再进行列截断
    df.columns = new_cols
    # 丢弃全是空值的行
    df.dropna(how='all', inplace=True)
    return df

def load_excel(file, start_row, row_count):
    try:
        file.seek(0)
        if file.name.lower().endswith('.xls'):
            df = pd.read_excel(file, header=list(range(start_row, start_row + row_count)), engine='xlrd')
        else:
            df = pd.read_excel(file, header=list(range(start_row, start_row + row_count)), engine='openpyxl')
        return clean_columns(df)
    except Exception as e:
        return None

def get_headers_only(file, start_row, row_count):
    try:
        file.seek(0)
        if file.name.lower().endswith('.xls'):
            df = pd.read_excel(file, header=list(range(start_row, start_row + row_count)), nrows=0, engine='xlrd')
        else:
            df = pd.read_excel(file, header=list(range(start_row, start_row + row_count)), nrows=0, engine='openpyxl')
        return clean_columns(df).columns.tolist()
    except: return []

def find_col_index(target, header_list):
    if not target: return -1
    try:
        # 直接匹配扁平化后的全名 (UI 看到什么，这里就匹配什么)
        return header_list.index(target) + 1
    except:
        # 如果全名没匹配上，尝试兼容性匹配 (去掉重名后缀)
        clean_t = target.rsplit('_', 1)[0] if '_' in target else target
        for i, h in enumerate(header_list):
            if h and clean_t in str(h): return i + 1
    return -1

def queue_download_js(results):
    import json
    files_list = []
    for name, data in results:
        b64 = base64.b64encode(data).decode()
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if name.lower().endswith('.xls'):
            mime = "application/vnd.ms-excel"
        files_list.append({"name": name, "base64": b64, "mime": mime})
    
    files_json = json.dumps(files_list)
    
    js_code = f"""
        <script>
        (function() {{
            const files = {files_json};
            async function download() {{
                for (const file of files) {{
                    const blob = await (await fetch(`data:${{file.mime}};base64,${{file.base64}}`)).blob();
                    const url = window.URL.createObjectURL(blob);
                    const link = document.createElement('a');
                    link.style.display = 'none';
                    link.href = url;
                    link.download = file.name;
                    document.body.appendChild(link);
                    link.click();
                    window.URL.revokeObjectURL(url);
                    document.body.removeChild(link);
                    await new Promise(r => setTimeout(r, 800));
                }}
            }}
            download();
        }})();
        </script>
    """
    return js_code

# --- 3. 界面交互 ---
st.markdown('<div class="main-title">⚡ 小雷 Excel 批量助手</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">高效、简单、本地处理的 Excel 办公利器</div>', unsafe_allow_html=True)

tab1, tab2 = st.tabs(["🔄 批量关联更新", "✨ 批量字段提取"])

with tab1:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    up_c1, up_c2 = st.columns([1.2, 2])
    with up_c1:
        st.markdown("##### � A表 (来源数据)")
        f_a = st.file_uploader("A", type=['xlsx', 'xls'], label_visibility="collapsed", key="ua")
        c_as, c_ac = st.columns(2)
        has = c_as.number_input("起行", 0, 20, 0, key="has_global")
        hac = c_ac.number_input("行数", 1, 5, 1, key="hac_global")
    with up_c2:
        st.markdown("##### � B表 (目标多选)")
        fs_b = st.file_uploader("B", type=['xlsx', 'xls'], accept_multiple_files=True, label_visibility="collapsed", key="ubs")
        c_bs, c_bc = st.columns(2)
        hbs = c_bs.number_input("起行", 0, 20, 0, key="hbs_global")
        hbc = c_bc.number_input("行数", 1, 5, 1, key="hbc_global")
    st.markdown('</div>', unsafe_allow_html=True)

    if f_a and fs_b:
        df_a_g = load_excel(f_a, has, hac)
        df_b_g = load_excel(fs_b[0], hbs, hbc)

        if df_a_g is not None and df_b_g is not None:
            st.markdown('<div class="config-card">', unsafe_allow_html=True)
            st.caption("⚙️ 全局默认配置 (修改后将同步至下方个性化默认值)")
            cols = st.columns(4)
            ak = cols[0].selectbox("A匹配列", df_a_g.columns, key="gak")
            av = cols[1].selectbox("A取值列", df_a_g.columns, key="gav")
            bk = cols[2].selectbox("B匹配列", df_b_g.columns, key="gbk")
            bt = cols[3].selectbox("B更新列", df_b_g.columns, key="gbt")
            st.markdown('</div>', unsafe_allow_html=True)

            overrides = {}
            with st.expander(f"🔍 个性化清单 ({len(fs_b)}份)"):
                for i, fb in enumerate(fs_b):
                    oc1, oc2 = st.columns([5, 1])
                    oc1.caption(f"📄 {fb.name}")
                    # 只要勾选了个性化，才记录 override
                    if oc2.checkbox("个性化", key=f"is_p_{i}"):
                        st.markdown('<div class="personal-box">', unsafe_allow_html=True)
                        r1, r2, r3, r4 = st.columns(4)
                        # 【核心修复】：value 直接绑定全局变量 has/hac/hbs/hbc
                        p_as = r1.number_input("A起", 0, 20, value=has, key=f"pas_{i}")
                        p_ac = r2.number_input("A行", 1, 5, value=hac, key=f"pac_{i}")
                        p_bs = r3.number_input("B起", 0, 20, value=hbs, key=f"pbs_{i}")
                        p_bc = r4.number_input("B行", 1, 5, value=hbc, key=f"pbc_{i}")
                        
                        df_al, df_bl = load_excel(f_a, p_as, p_ac), load_excel(fb, p_bs, p_bc)
                        if df_al is not None and df_bl is not None:
                            def get_safe_idx(options, target):
                                try: return list(options).index(target)
                                except: return 0
                            l1, l2, l3, l4 = st.columns(4)
                            # 【核心修复】：index 动态根据全局选中的列名计算
                            p_ak = l1.selectbox("A匹", df_al.columns, index=get_safe_idx(df_al.columns, ak), key=f"pak_{i}")
                            p_av = l2.selectbox("A取", df_al.columns, index=get_safe_idx(df_al.columns, av), key=f"pav_{i}")
                            p_bk = l3.selectbox("B匹", df_bl.columns, index=get_safe_idx(df_bl.columns, bk), key=f"pbk_{i}")
                            p_bt = l4.selectbox("B更", df_bl.columns, index=get_safe_idx(df_bl.columns, bt), key=f"pbt_{i}")
                            overrides[i] = {'has':p_as,'hac':p_ac,'hbs':p_bs,'hbc':p_bc,'ak':p_ak,'av':p_av,'bk':p_bk,'bt':p_bt}
                        st.markdown('</div>', unsafe_allow_html=True)

            if st.button("🚀 开始批量处理", use_container_width=True):
                temp_results = []
                error_logs = []  # 新增：记录错误信息
                prog = st.progress(0)
                for i, fb in enumerate(fs_b):
                    try:
                        # 逻辑：有个性化配置用个性化，没有用全局
                        conf = overrides.get(i, {'has':has,'hac':hac,'hbs':hbs,'hbc':hbc,'ak':ak,'av':av,'bk':bk,'bt':bt})
                        df_ca = load_excel(f_a, conf['has'], conf['hac'])
                        mapping = dict(zip(df_ca[conf['ak']].astype(str).str.strip(), df_ca[conf['av']]))
                        
                        # 获取扁平化后的表头列表，确保与 UI 看到的名称完全一致
                        headers = get_headers_only(fb, conf['hbs'], conf['hbc'])
                        ik, it = find_col_index(conf['bk'], headers), find_col_index(conf['bt'], headers)
                        
                        if ik != -1 and it != -1:
                            fb.seek(0)
                            file_ext = fb.name.rsplit('.', 1)[-1].lower()
                            
                            if file_ext == 'xls':
                                # .xls 使用 xlrd + xlutils 以保留格式
                                rb = xlrd.open_workbook(file_contents=fb.read(), formatting_info=True)
                                wb = xl_copy(rb)
                                ws = wb.get_sheet(0)
                                sheet_read = rb.sheet_by_index(0)
                                
                                col_k_idx = ik - 1
                                col_t_idx = it - 1
                                start_row_idx = conf['hbs'] + conf['hbc']
                                
                                for r in range(start_row_idx, sheet_read.nrows):
                                    cell_val = sheet_read.cell_value(r, col_k_idx)
                                    kv = str(cell_val).strip()
                                    if kv in mapping:
                                        ws.write(r, col_t_idx, mapping[kv])
                                
                                out = BytesIO()
                                wb.save(out)
                                temp_results.append((fb.name, out.getvalue()))
                            else:
                                # .xlsx 使用 openpyxl
                                wb = openpyxl.load_workbook(fb)
                                ws = wb.active
                                h_row = conf['hbs'] + conf['hbc']
                                for r in range(h_row + 1, ws.max_row + 1):
                                    kv = str(ws.cell(r, ik).value or "").strip()
                                    if kv in mapping: ws.cell(r, it).value = mapping[kv]
                                out = BytesIO(); wb.save(out)
                                temp_results.append((fb.name.rsplit('.', 1)[0] + ".xlsx", out.getvalue()))
                        else:
                            # 记录未匹配到列的错误
                            missing = []
                            if ik == -1: missing.append(f"匹配列 '{conf['bk']}'")
                            if it == -1: missing.append(f"更新列 '{conf['bt']}'")
                            error_logs.append(f"❌ {fb.name}: 未找到 {' 和 '.join(missing)}")
                    except Exception as e:
                        error_logs.append(f"⚠️ {fb.name}: 处理出错 - {str(e)}")
                    prog.progress((i + 1) / len(fs_b))
                
                st.session_state.batch_results = temp_results
                
                # 如果有错误，显示出来
                if error_logs:
                    with st.expander("🚨 处理异常报告", expanded=True):
                        for log in error_logs:
                            st.write(log)

            if st.session_state.batch_results:
                res = st.session_state.batch_results
                st.markdown(f'<div class="download-bar"><span>✅ 处理完成 ({len(res)}个)</span></div>', unsafe_allow_html=True)
                if st.button(f"📥 按顺序自动下载全部 {len(res)} 个文件", use_container_width=True, type="primary"):
                    # 使用 key 强制刷新组件以触发 JS
                    st.components.v1.html(queue_download_js(res), height=0)
                    st.toast("正在启动队列下载，请允许浏览器下载多个文件...", icon="⌛")
                    # --- TAB 2: 字段提取 (重写加回) ---
with tab2:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    st.markdown("##### 🔍 批量提取数据字段")
    fs_ex = st.file_uploader("提取文件", type=['xlsx', 'xls'], accept_multiple_files=True, label_visibility="collapsed", key="uex")
    if fs_ex:
        st.markdown('<div class="config-card">', unsafe_allow_html=True)
        st.markdown("##### ⚙️ 全局默认设置")
        c1, c2 = st.columns(2)
        exs_g = c1.number_input("默认表头起行", 0, 10, 0, key="exs_g")
        exc_g = c2.number_input("默认表头行数", 1, 5, 1, key="exc_g")
        
        sample_df_g = load_excel(fs_ex[0], exs_g, exc_g)
        sel_cols_g = []
        if sample_df_g is not None:
            sel_cols_g = st.multiselect("默认勾选字段 (将作为所有文件的初始选择)：", options=sample_df_g.columns, key="sel_cols_g")
        st.markdown('</div>', unsafe_allow_html=True)

        # 记录每个文件的个性化配置
        ex_overrides = {}
        with st.expander(f"📋 个性化提取清单 ({len(fs_ex)}份)", expanded=False):
            for i, f in enumerate(fs_ex):
                st.markdown(f'<div class="personal-box">', unsafe_allow_html=True)
                st.markdown(f"📄 **{f.name}**")
                ec1, ec2 = st.columns(2)
                f_exs = ec1.number_input("起行", 0, 10, value=exs_g, key=f"exs_{i}")
                f_exc = ec2.number_input("行数", 1, 5, value=exc_g, key=f"exc_{i}")
                
                f_df = load_excel(f, f_exs, f_exc)
                if f_df is not None:
                    # 默认选中在全局配置中存在的列
                    default_sel = [c for c in sel_cols_g if c in f_df.columns]
                    f_sel = st.multiselect("选择字段：", options=f_df.columns, default=default_sel, key=f"sel_{i}")
                    ex_overrides[i] = {'exs': f_exs, 'exc': f_exc, 'sel': f_sel}
                st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="config-card">', unsafe_allow_html=True)
        st.markdown("##### 📝 导出配置")
        c_n1, c_n2 = st.columns([3, 1])
        default_name_tpl = "提取名"
        ex_name_tpl = c_n1.text_input("导出文件名 (使用 {name} 代替原文件名)", value=default_name_tpl, key="ex_name_tpl")
        merge_all = c_n2.checkbox("合并为一个表", value=False, key="merge_ex")
        st.markdown('</div>', unsafe_allow_html=True)

        if st.button("🚀 执行批量提取", use_container_width=True, key="run_ex"):
            ex_results = []
            all_dfs = [] 
            ex_error_logs = [] 
            prog_ex = st.progress(0)
            
            # 如果是合并模式，先确定主列列表
            master_cols = sel_cols_g if merge_all else []
            
            for i, f in enumerate(fs_ex):
                try:
                    conf = ex_overrides.get(i, {'exs': exs_g, 'exc': exc_g, 'sel': sel_cols_g})
                    current_sel = conf['sel']
                    
                    if not current_sel and not merge_all:
                        continue

                    if merge_all:
                        df = load_excel(f, conf['exs'], conf['exc'])
                        if df is not None and not df.empty:
                            target_cols = current_sel if current_sel else master_cols
                            # 建立该文件的列映射（忽略后缀匹配以增加鲁棒性）
                            file_col_map = {}
                            for col in df.columns:
                                base = col.rsplit('_', 1)[0] if '_' in col else col
                                if base not in file_col_map: file_col_map[base] = col
                            
                            # 按照选中的列顺序提取数据
                            extracted_data = {}
                            found_any = False
                            for t_col in target_cols:
                                if t_col in df.columns:
                                    extracted_data[t_col] = df[t_col]
                                    found_any = True
                                else:
                                    # 尝试忽略后缀匹配
                                    t_base = t_col.rsplit('_', 1)[0] if '_' in t_col else t_col
                                    if t_base in file_col_map:
                                        extracted_data[t_col] = df[file_col_map[t_base]]
                                        found_any = True
                                    else:
                                        extracted_data[t_col] = pd.Series([None] * len(df))
                            
                            if found_any:
                                # 【核心修复】：将提取的数据转换为标准的 DataFrame
                                # 使用字典构造，并明确指定每一列的数据内容
                                df_to_append = pd.DataFrame()
                                for t_col in target_cols:
                                    # 如果该列没匹配到，填充等长的空值
                                    data = extracted_data.get(t_col)
                                    if data is None:
                                        data = [None] * len(df)
                                    df_to_append[t_col] = data
                                
                                # 重置索引并加入列表
                                df_to_append.reset_index(drop=True, inplace=True)
                                all_dfs.append(df_to_append)
                                ex_error_logs.append(f"✅ {f.name}: 提取并对齐了 {len(df_to_append)} 行数据")
                            else:
                                ex_error_logs.append(f"⚠️ {f.name}: 未匹配到任何选中的列")
                        else:
                            ex_error_logs.append(f"❌ {f.name}: 读取失败或为空")
                    else:
                        # 非合并模式：使用 openpyxl 保持原样逻辑
                        f.seek(0)
                        if f.name.lower().endswith('.xls'):
                            stream = to_xlsx_stream(f)
                        else:
                            stream = f
                        
                        if stream is None:
                            ex_error_logs.append(f"❌ {f.name}: .xls 转换失败")
                            continue
                            
                        wb = openpyxl.load_workbook(stream)
                        ws = wb.active
                        file_headers = get_headers_only(f, conf['exs'], conf['exc'])
                        
                        cols_to_delete = []
                        for idx, h in enumerate(file_headers):
                            if h not in current_sel:
                                cols_to_delete.append(idx + 1)
                        
                        if len(cols_to_delete) == len(file_headers):
                            ex_error_logs.append(f"⚠️ {f.name}: 未选中任何有效列")
                            continue

                        for col_idx in sorted(cols_to_delete, reverse=True):
                            ws.delete_cols(col_idx)
                        
                        out = BytesIO()
                        wb.save(out)
                        orig_name = f.name.rsplit('.', 1)[0]
                        new_name_base = ex_name_tpl.replace("{name}", orig_name) if "{name}" in ex_name_tpl else (f"{ex_name_tpl}_{i+1}" if len(fs_ex) > 1 else ex_name_tpl)
                        ex_results.append((f"{new_name_base}.xlsx", out.getvalue()))
                        ex_error_logs.append(f"✅ {f.name}: 已提取并保留格式")
                except Exception as e:
                    ex_error_logs.append(f"⚠️ {f.name}: 处理出错 - {str(e)}")
                prog_ex.progress((i + 1) / len(fs_ex))
            
            if merge_all and all_dfs:
                # 智能合并：对齐列并忽略索引
                merged_df = pd.concat(all_dfs, ignore_index=True)
                out = BytesIO()
                with pd.ExcelWriter(out, engine='openpyxl') as writer:
                    merged_df.to_excel(writer, index=False)
                ex_results = [(f"{ex_name_tpl}.xlsx", out.getvalue())]
                st.success(f"📊 汇总完成：共计 {len(merged_df)} 行数据")
            
            st.session_state.extract_results = ex_results

            # 显示错误报告
            if ex_error_logs:
                with st.expander("🚨 提取异常报告", expanded=True):
                    for log in ex_error_logs:
                        st.write(log)

            if st.session_state.extract_results:
                e_res = st.session_state.extract_results
                st.markdown(f'<div class="download-bar"><span>✅ 提取完成 ({len(e_res)}个)</span></div>', unsafe_allow_html=True)
                if st.button(f"📥 顺序自动下载 {len(e_res)} 个提取文件", use_container_width=True, type="primary", key="dl_ex_btn"):
                    # 使用当前时间戳作为 key 强制刷新 HTML 组件以触发 JS 执行
                    st.components.v1.html(queue_download_js(e_res), height=0)
                    st.toast("正在启动队列下载，请允许浏览器下载多个文件...", icon="⌛")
                
                with st.expander("📄 查看提取文件明细"):
                    for idx, (n, d) in enumerate(e_res):
                        cl1, cl2 = st.columns([4, 1])
                        cl1.markdown(f"📄 {n}")
                        cl2.download_button("下载", d, n, key=f"ex_dl_{idx}")

# --- 4. 底部声明 ---
st.markdown("---")
st.markdown(
    '<div style="text-align: center; color: #94a3b8; font-size: 0.8rem; padding: 20px;">'
    '小雷 Excel 助手 · 本地处理更安全 · 2024'
    '</div>', 
    unsafe_allow_html=True
)